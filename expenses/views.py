from django.shortcuts import render
from .models import Expense

def expense_list(request):
    expenses = Expense.objects.filter(user=request.user)
    return render(request, 'expenses/expense_list.html', {'expenses': expenses})

from django.shortcuts import render, redirect
from .forms import ExpenseForm

def add_expense(request):
    if request.method == 'POST':
        form = ExpenseForm(request.POST)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.user = request.user
            expense.save()
            return redirect('expense_list')
    else:
        form = ExpenseForm()
    return render(request, 'expenses/add_expense.html', {'form': form})

from django.contrib.auth.decorators import login_required

@login_required
def expense_list(request):
    ...
from django.http import HttpResponse

def index(request):
    return HttpResponse("Hello, this is the expenses index page.")

import openpyxl
from django.http import HttpResponse
from .models import Expense

def export_expenses_xls(request):
    # Create a workbook and an active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Expenses'

    # Define the headers
    headers = ['Date', 'Category', 'Amount', 'Description']

    # Write the headers to the first row
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header

    # Get all expenses from the database
    expenses = Expense.objects.filter(user=request.user)

    # Write the data to the worksheet
    for row_num, expense in enumerate(expenses, 2):
        worksheet.cell(row=row_num, column=1).value = expense.date
        worksheet.cell(row=row_num, column=2).value = expense.category
        worksheet.cell(row=row_num, column=3).value = expense.amount
        worksheet.cell(row=row_num, column=4).value = expense.description

    # Prepare the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=expenses.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response
